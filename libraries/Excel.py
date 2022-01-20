
from builtins import len
from typing import Any
from io import BytesIO
from typing import Any, Dict, Iterator, List, Optional, Tuple

import pandas as pd
import xlrd
import datetime

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from customlibs.xlreader import ExcelRead

from customlibs import xlreader


from openpyxl import load_workbook
from robot.libraries.BuiltIn import BuiltIn

class SuchIdIsExistException(Exception):
    """Raised when the document with the identifier is already in the cache."""
    pass


class NoSuchIdException(Exception):
    """Raised when accessing an absent document identifier."""
    pass


class NoOpenedDocumentsException(Exception):
    """Raised in the absence of open documents."""
    pass

class Excel:
    """This library was developed for easy data management using Excel documents, giving the user a easier
    solution to edit execution data on test projects.

    *Copyright 2020-*     NEORIS | Jesus Barajas \n
    *Update 2020, July 1st*		NEORIS | Israel Alvarez

    = Table of contents =

    - `Requirements`
    - `Excel sheet Data Format`
    - `Usage`
    - `Examples`

    = Requirements =
        As described in the requirements section, you have to install the latest versions of the described
        elements in this section, the following sections describe how to write Excel formated files
        to use the library and explain the process of import of this single keyword library.
        - RobotFramework installation.
        - Python 2.7.15 or above.
        - Pandas Library for Python.
        - Openpyxl Library for Python

    = Excel sheet Data Format =
    Data files can be read in 2 formats:

    Format 1 :
    Data files has to be written in a specific format, the two first cells on the top are taken as place holders
    Note:"Use this syntax in order to make the library work"

    This is an example of a correct syntax of a document to load the variables with the given name.

    | Variable | DATA |
    | URL | www.RobotFramework.com |
    | user | sampleuser123 |
    | pass | secret123 |

    Format 2 :
    Data file can be written as rows where the first row is taken as HEADERS and the following rows are the values
    and we could select a specific row.

    | Hdr1 | Hdr2 | Hdr... |
    | Row1 | Val | Val... |
    | Row2 | Val | Val... |


    = Usage =

    This library has 2 keyword and can be used depends on the desire approach.
    To use the library you should import the *Excel.py* file into your project, then you will be able to access the
    keywords *Read Data* and *Read Excel By Row*

    = Examples =

    *Example 1 : Read Data*

    This is an example where we want to build our global variables in the environment.
    First we need to write our Excel document with the format described in the *Excel sheet Format section.*

    | Variable | DATA |
    | Global | This is a global variable |
    | Global2 | This is a global variable too |
    | Global3 | This one is a global variables |

    Remember to save this file with the according name as described in *Excel sheet Format section.*
        - T=Test Level Variables, *Example*: T-Variables1.xlxs
        - G=Global Level Variables *Example*: G-Variables2.xlxs
        - S=Suite Level Variables *Example*: S-Variables3.xlxs

    *Example 2 : Read Excel By Row*

    This is an example to build our global variables in the environment from excel data.

    | Iteration | Browser | URL | USER | PASS |
    | 1 | chrome | www.RobotFramework.org | admin | 1234 |
    | 2 | edge | www.python.org | guest | 1234 |
    | ... | ... | ... | ... | ... |


    """
    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'

    __version__ = '2.0'

    def __init__(self):
        self.read = None
        self.wb = None
        self.ws = None
        self.wb2 = None
        self.ws2 = None
        self.dict_row = None
        self._cache: Dict[str, openpyxl.Workbook] = {}
        self._current_id: Optional[str] = None



    def get_column_number_by_header(filename, sheet_name, header):
        read = ExcelRead(filename)
        wb = xlrd.open_workbook(read.file_path)
        ws = wb.sheet_by_name(sheet_name)
        ws.cell_value(0, 0)
        for i in range(ws.ncols):
            value = ws.cell_value(0, i)
            if value == header:
                break
            elif value != header and ws.cell_value(0, i + 1) is None:
                raise IOError('Header not found')
        wb.release_resources()
        column = i + 1
        wb.release_resources()
        return column


    def read_excel_by_row(filename, row, sheet_name):
        """
            This Keyword receive 2 required arguments & 1 optional to create Test Variables in RobotFramework.

            *Example: *

            | Read Excel By Row | PathToDataFile | Desired Row | optional Sheet name |
            | Read Excel By Row | ${EXECDIR}\\TestData\\filename | 4 | Test |

        """
        wb = load_workbook(filename, read_only=True, data_only=True, keep_links=True)
        ws = wb[sheet_name]
        hdrs = []
        values = []
        x = 0
        r = int(row)
        for item in ws.values:
            if x == 0:
                hdrs = list(item)
            elif x == r:
                values = list(item)
                break
            x += 1
        dict_row = dict(zip(hdrs, values))
        wb.close()
        return dict_row
    def write_excel_cell(self, row_num: int, col_num: int, value: Any, sheet_name: str = None) -> None:
        """Writes value to the cell.\n
        *Args:*\n
            _row_num_: row number, starts with 1.\n
            _col_num_: column number, starts with 1.\n
            _value_: value for writing to a cell.\n
            _sheet_name_: sheet name for write.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | doc_id=docname1 |
        | Write Excel Cell | row_num=1 | col_num=3 | value=a3 | sheet_name=${DEFAULT_SHEET_NAME} |
        | Close All Excel Documents |
        """
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row=row_num, column=col_num, value=value)

    def write_excel_by_cell(filename, sheet_name, row_num, col_num, value: Any) -> None:
        """Writes value to the cell.\n
        Args:\n
            row_num: row number, starts with 1.\n
            col_num: column number, starts with 1.\n
            value: value for writing to a cell.\n
            sheet_name: sheet name for write.\n
        Example:\n
        | Write Excel Cell | row_num=1 | col_num=3 | value=a3 | sheet_name=${DEFAULT_SHEET_NAME} |
        """
        read = ExcelRead(filename)
        wb = load_workbook(filename, data_only=True, keep_links=True)
        ws = wb[sheet_name]
        ws.cell(row_num, col_num).value = value
        wb.save(read.file_path)
        wb.close()

    def load_data_table(self, filename, sheet, row):
        #file_name = "D:\\Automation\\Proyectos\\POCS\\DEMO\\TestData\\G-Variables2.xlsx"
        self.file_name = filename
        self.sheet = sheet
        self.row = row
        self.data = Excel.read_excel_by_row(self.file_name, row=row, sheet_name=self.sheet)
        exec_time = datetime.datetime.now()
        #Keywords.write_output(self, col_name='startExec', value=exec_time)
        '''for k, v in self.data.items():
            print(f'{k} : {v}')
        print('***\n')'''
        return self.data

    def open_excel_document(self, filename: str, doc_id: str) -> str:
        """Opens xlsx document file.\n
        *Args:*\n
            _filename_: document name.\n
            _doc_id_: the identifier for the document that will be opened.\n
        *Returns:*\n
            Document identifier from the cache.\n
        *Example:*\n
        | Open Excel Document | filename=file.xlsx | doc_id=docid |
        | Close All Excel Documents |
        """
        filename = str(filename)
        doc_id = str(doc_id)
        if doc_id in self._cache:
            message = u"Document with such id {0} is opened."
            raise SuchIdIsExistException(message.format(doc_id))
        workbook = openpyxl.load_workbook(filename=filename)
        self._cache[doc_id] = workbook
        self._current_id = doc_id
        return self._current_id

    def _get_current_workbook(self) -> openpyxl.Workbook:
        """Checks opened document.\n
        *Returns:*\n
            Current document.\n
        """
        if not self._cache or not self._current_id:
            raise NoOpenedDocumentsException(u"No opened documents in cache.")
        return self._cache[self._current_id]

    def get_sheet(self, sheet_name: str = None) -> Worksheet:
        """Returns a page from the current document.\n
        *Args:*\n
            _sheet_name_: sheet name.\n
        *Returns:*\n
            Document's sheet.\n
        """
        workbook = self._get_current_workbook()
        if sheet_name is None:
            return workbook.active
        sheet_name = str(sheet_name)
        return workbook[sheet_name]


    def count_row(self, sheet_name=None):
        """Getting row count. \n"""
        sheet = self.get_sheet(sheet_name)
        rcount = sheet.max_row
        return rcount

    def save_excel_document(self, filename: str) -> None:
        """Saves the current document to disk.\n
        *Args:*\n
            _filename_: file name to save.\n
        *Example:*\n
        | Create Excel Document | doc_id=doc1 |
        | Save Excel Document | filename=file1.xlsx |
        | Close All Excel Documents |
        """
        workbook = self._get_current_workbook()
        workbook.save(filename=filename)

    def close_current_excel_document(self) -> Optional[str]:
        """Closes current document.\n
        *Returns:*\n
            Closed document identifier.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | docname1 |
        | ${doc2}= | Create Excel Document | docname2 |
        | Close Current Excel Document |
        """
        if self._current_id is not None:
            self._cache.pop(self._current_id)
            self._current_id = None
        if self._cache:
            self._current_id = list(self._cache.keys())[0]
        return self._current_id
